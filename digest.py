import os
import json
import re
import smtplib
import requests
import xml.etree.ElementTree as ET
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import pandas as pd
import google.generativeai as genai
import gspread
from google.oauth2.service_account import Credentials
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Setup directories and configs
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH = os.path.join(BASE_DIR, "config.json")
SEEN_JOBS_PATH = os.path.join(BASE_DIR, "seen_jobs.json")

# Standard headers to bypass basic anti-bot blocks
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.5",
    "Referer": "https://www.google.com/"
}

def load_config():
    if not os.path.exists(CONFIG_PATH):
        raise FileNotFoundError(f"Configuration file not found at {CONFIG_PATH}")
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        return json.load(f)

def load_seen_jobs():
    if os.path.exists(SEEN_JOBS_PATH):
        try:
            with open(SEEN_JOBS_PATH, "r", encoding="utf-8") as f:
                return set(json.load(f))
        except Exception:
            return set()
    return set()

def save_seen_jobs(seen_set):
    with open(SEEN_JOBS_PATH, "w", encoding="utf-8") as f:
        json.dump(list(seen_set), f, indent=2)

def clean_html(raw_html):
    if not raw_html:
        return ""
    # Strip HTML tags
    clean = re.sub(r'<[^>]*>', ' ', raw_html)
    # Decode entities
    clean = clean.replace("&nbsp;", " ").replace("&amp;", "&").replace("&lt;", "<").replace("&gt;", ">").replace("&quot;", '"')
    return " ".join(clean.split()).strip()

def is_url_alive(url):
    if not url or not url.startswith("http"):
        return False
    try:
        # 1. Attempt quick HEAD check
        res = requests.head(url, headers=HEADERS, timeout=3.0, allow_redirects=True)
        if res.status_code == 404:
            return False
        if 200 <= res.status_code < 400:
            return True

        # 2. Attempt fallback GET check
        get_res = requests.get(url, headers=HEADERS, timeout=3.5, allow_redirects=True)
        if get_res.status_code == 404:
            return False
        if 200 <= get_res.status_code < 400:
            return True

        # If rate-limited or challenged (403, 429, 503, >=500), assume it's alive for human browsers
        if get_res.status_code in [403, 429, 503] or get_res.status_code >= 500:
            return True
        return False
    except requests.exceptions.Timeout:
        # Slow sites are likely alive for humans
        return True
    except Exception:
        return False

# ================= HTTP REQUEST UTILITY =================

def make_request(url, params=None, headers=None):
    req_headers = {"User-Agent": "Mozilla/5.0 compatible"}
    if headers:
        req_headers.update(headers)
        
    for attempt in range(2):
        try:
            res = requests.get(url, params=params, headers=req_headers, timeout=10.0)
            if res.status_code == 200:
                return res
            # If 404, don't retry
            if res.status_code == 404:
                return None
        except Exception as e:
            if attempt == 0:
                print(f"[Request] Attempt 1 failed for {url}: {e}. Retrying...")
            else:
                print(f"[Request] Attempt 2 failed for {url}: {e}.")
    return None

# ================= SCRAPERS =================

def scrape_remotive():
    jobs = []
    try:
        print("[Scraper] Querying Remotive API...")
        res = make_request("https://remotive.com/api/remote-jobs?category=devops-sysadmin")
        if res:
            data = res.json()
            for r_job in data.get("jobs", []):
                jobs.append({
                    "title": r_job.get("title", ""),
                    "company": r_job.get("company_name", ""),
                    "location": r_job.get("candidate_required_location", "Remote"),
                    "is_remote": True,
                    "source": "remotive",
                    "url": r_job.get("url", ""),
                    "description": clean_html(r_job.get("description", "")),
                    "posted_date": r_job.get("publication_date", datetime.now().isoformat())
                })
    except Exception as e:
        print(f"[Scraper Error] Remotive failed: {e}")
    return jobs

def scrape_jobicy():
    jobs = []
    try:
        print("[Scraper] Querying Jobicy API...")
        res = make_request("https://jobicy.com/api/v2/remote-jobs?count=50&industry=engineering")
        if res:
            data = res.json()
            for j_job in data.get("jobs", []):
                jobs.append({
                    "title": j_job.get("jobTitle", ""),
                    "company": j_job.get("companyName", ""),
                    "location": j_job.get("jobGeo", "Remote"),
                    "is_remote": True,
                    "source": "jobicy",
                    "url": j_job.get("jobUrl", ""),
                    "description": clean_html(j_job.get("jobDescription", "")),
                    "posted_date": j_job.get("pubDate", datetime.now().isoformat())
                })
    except Exception as e:
        print(f"[Scraper Error] Jobicy failed: {e}")
    return jobs

def scrape_wwr():
    jobs = []
    try:
        print("[Scraper] Fetching We Work Remotely RSS...")
        import feedparser
        res = make_request("https://weworkremotely.com/categories/remote-devops-sysadmin-jobs.rss")
        if res:
            feed = feedparser.parse(res.text)
            for entry in feed.entries:
                title_text = entry.get("title", "")
                url = entry.get("link", "")
                desc = clean_html(entry.get("summary", ""))
                posted = entry.get("published", datetime.now().isoformat())
                
                title = title_text
                company = "We Work Remotely"
                if ":" in title_text:
                     parts = title_text.split(":", 1)
                     company = parts[0].strip()
                     title = parts[1].strip()
                elif " at " in title_text:
                     parts = title_text.split(" at ", 1)
                     title = parts[0].strip()
                     company = parts[1].strip()

                jobs.append({
                    "title": title,
                    "company": company,
                    "location": "Remote",
                    "is_remote": True,
                    "source": "weworkremotely",
                    "url": url,
                    "description": desc,
                    "posted_date": posted
                })
    except Exception as e:
        print(f"[Scraper Error] WWR failed: {e}")
    return jobs

def scrape_otta():
    jobs = []
    try:
        print("[Scraper] Querying Otta Public Search API...")
        res = make_request("https://app.otta.com/api/jobs?role=devops&location=india")
        if res:
            data = res.json()
            jobs_list = []
            if isinstance(data, list):
                jobs_list = data
            elif isinstance(data, dict):
                if "jobs" in data:
                    jobs_list = data["jobs"]
                elif "results" in data:
                    jobs_list = data["results"]
                elif "data" in data:
                    jobs_list = data["data"]
            
            for item in jobs_list:
                title = item.get("title") or item.get("jobTitle") or ""
                company = ""
                if "company" in item:
                    if isinstance(item["company"], dict):
                        company = item["company"].get("name", "")
                    else:
                        company = str(item["company"])
                else:
                    company = item.get("companyName") or item.get("company_name") or "Otta Employer"
                    
                url = item.get("url") or item.get("jobUrl") or item.get("link") or ""
                loc = item.get("location") or item.get("jobLocation") or "Remote / India"
                desc = clean_html(item.get("description") or item.get("jobDescription") or item.get("summary") or "")
                posted = item.get("posted_date") or item.get("created_at") or datetime.now().isoformat()
                
                if title and url:
                    jobs.append({
                        "title": title,
                        "company": company,
                        "location": loc,
                        "is_remote": "remote" in loc.lower() or loc.lower() == "remote",
                        "source": "otta",
                        "url": url,
                        "description": desc,
                        "posted_date": posted
                    })
    except Exception as e:
        print(f"[Scraper Error] Otta failed: {e}")
    return jobs

def scrape_arbeitnow():
    jobs = []
    try:
        print("[Scraper] Querying Arbeitnow API...")
        res = make_request("https://www.arbeitnow.com/api/job-board-api?tags[]=devops&tags[]=kubernetes")
        if res:
            data = res.json()
            for item in data.get("data", []):
                title = item.get("title", "")
                company = item.get("company_name", "")
                url = item.get("url", "")
                loc = item.get("location", "Remote")
                desc = clean_html(item.get("description", ""))
                posted = item.get("created_at", datetime.now().isoformat())
                
                jobs.append({
                    "title": title,
                    "company": company,
                    "location": loc,
                    "is_remote": "remote" in loc.lower() or item.get("remote", False),
                    "source": "arbeitnow",
                    "url": url,
                    "description": desc,
                    "posted_date": posted
                })
    except Exception as e:
        print(f"[Scraper Error] Arbeitnow failed: {e}")
    return jobs

def scrape_naukri():
    jobs = []
    try:
        print("[Scraper] Querying Naukri API...")
        res = make_request("https://www.naukri.com/jobapi/v3/search?noOfResults=20&urlType=search_by_keyword&searchType=adv&keyword=devops+kubernetes&location=hyderabad%2Cbangalore")
        if not res:
            print("[Scraper Warning] Naukri API was blocked or returned an error. Skipping.")
            return []
        
        data = res.json()
        search_results = []
        if isinstance(data, dict):
            search_results = data.get("searchResults", []) or data.get("jobs", []) or data.get("data", [])
        elif isinstance(data, list):
            search_results = data
            
        for item in search_results:
            title = item.get("title") or item.get("jobTitle") or ""
            company = ""
            if "companyName" in item:
                company = item["companyName"]
            elif "company" in item:
                if isinstance(item["company"], dict):
                    company = item["company"].get("name", "")
                else:
                    company = str(item["company"])
            else:
                company = "Naukri Employer"
                
            url = item.get("jdURL") or item.get("jobUrl") or item.get("url") or item.get("link") or ""
            loc = item.get("place") or item.get("location") or "Hyderabad/Bangalore"
            desc = clean_html(item.get("description") or item.get("jobDescription") or item.get("summary") or "")
            posted = item.get("postedDate") or item.get("createdDate") or datetime.now().isoformat()
            
            if title and url:
                jobs.append({
                    "title": title,
                    "company": company,
                    "location": loc,
                    "is_remote": "remote" in loc.lower(),
                    "source": "naukri",
                    "url": url,
                    "description": desc,
                    "posted_date": posted
                })
    except Exception as e:
        print(f"[Scraper Warning] Failed to parse Naukri response: {e}. Skipping.")
    return jobs

def scrape_instahyre():
    jobs = []
    try:
        print("[Scraper] Querying Instahyre API...")
        res = make_request("https://www.instahyre.com/api/v1/opportunity/?format=json&search=devops")
        if res:
            data = res.json()
            opportunities = data.get("opportunities", []) or data.get("results", []) or []
            for item in opportunities:
                title = item.get("title", "")
                company = "Instahyre Employer"
                if "company" in item:
                    if isinstance(item["company"], dict):
                        company = item["company"].get("name", "Instahyre Employer")
                    else:
                        company = str(item["company"])
                        
                loc = item.get("location", "India / Remote")
                if isinstance(loc, list):
                    loc = ", ".join(loc)
                    
                desc = clean_html(item.get("description") or item.get("job_description") or "")
                
                url = item.get("url") or item.get("link") or ""
                if not url and "id" in item:
                    url = f"https://www.instahyre.com/jobs/{item['id']}/"
                    
                posted = item.get("posted_date") or item.get("created_at") or datetime.now().isoformat()
                
                if title and url:
                    jobs.append({
                        "title": title,
                        "company": company,
                        "location": loc,
                        "is_remote": "remote" in loc.lower(),
                        "source": "instahyre",
                        "url": url,
                        "description": desc,
                        "posted_date": posted
                    })
    except Exception as e:
        print(f"[Scraper Error] Instahyre failed: {e}")
    return jobs


def scrape_jobspy():
    jobs = []
    try:
        print("[Scraper] Importing python-jobspy...")
        from jobspy import scrape_jobs
        
        # We query for both devops and sre in India / Remote
        queries = [
            {"search_term": "devops", "location": "India", "country_indeed": "India"},
            {"search_term": "sre", "location": "India", "country_indeed": "India"}
        ]
        
        for q in queries:
            try:
                print(f"[Scraper] Querying JobSpy for '{q['search_term']}' in '{q['location']}'...")
                res_df = scrape_jobs(
                    site_name=["linkedin", "indeed", "zip_recruiter"],
                    search_term=q["search_term"],
                    location=q["location"],
                    country_indeed=q["country_indeed"],
                    results_wanted=15
                )
                
                if res_df is None or res_df.empty:
                    continue
                
                # Standardize columns to lowercase
                res_df.columns = [c.lower() for c in res_df.columns]
                res_df = res_df.fillna("")
                
                for _, row in res_df.iterrows():
                    title = row.get("title", "")
                    company = row.get("company", "JobSpy Employer")
                    location = row.get("location", "India")
                    
                    site = row.get("site", "jobspy")
                    source = f"jobspy-{site}"
                    
                    url = row.get("job_url_direct") or row.get("job_url") or row.get("url") or ""
                    desc = clean_html(row.get("description", ""))
                    
                    is_remote = False
                    if "is_remote" in row:
                        is_remote = bool(row["is_remote"])
                    else:
                        is_remote = "remote" in str(location).lower()
                        
                    posted = row.get("date_posted") or row.get("posted_date") or datetime.now().isoformat()
                    if hasattr(posted, "isoformat"):
                        posted = posted.isoformat()
                    else:
                        posted = str(posted)
                        
                    if title and url:
                        jobs.append({
                            "title": str(title),
                            "company": str(company),
                            "location": str(location),
                            "is_remote": is_remote,
                            "source": source,
                            "url": str(url),
                            "description": desc,
                            "posted_date": posted
                        })
            except Exception as e:
                print(f"[Scraper Warning] JobSpy query for '{q['search_term']}' failed: {e}")
                
    except ImportError:
        print("[Scraper Warning] python-jobspy not installed. Skipping JobSpy (LinkedIn, Indeed, Google Jobs) sources.")
        print("[Scraper Warning] To enable, run: pip install python-jobspy")
    except Exception as e:
        print(f"[Scraper Error] JobSpy scraping failed: {e}")
    return jobs

def scrape_remoteok():
    jobs = []
    try:
        print("[Scraper] Querying RemoteOK API...")
        res = make_request("https://remoteok.com/api")
        if res:
            data = res.json()
            if isinstance(data, list) and len(data) > 1:
                for item in data[1:]:
                    if not isinstance(item, dict):
                        continue
                    if "legal" in item:
                        continue
                    
                    title = item.get("position", "")
                    company = item.get("company", "")
                    url = item.get("url", "")
                    desc = clean_html(item.get("description", ""))
                    posted = item.get("date", datetime.now().isoformat())
                    loc = item.get("location") or "Remote"
                    
                    if title and url:
                        jobs.append({
                            "title": title,
                            "company": company,
                            "location": loc,
                            "is_remote": True,
                            "source": "remoteok",
                            "url": url,
                            "description": desc,
                            "posted_date": posted
                        })
    except Exception as e:
        print(f"[Scraper Error] RemoteOK failed: {e}")
    return jobs

def scrape_hn_jobs():
    jobs = []
    try:
        print("[Scraper] Fetching Hacker News Jobs RSS...")
        import feedparser
        res = make_request("https://news.ycombinator.com/jobsrss")
        if res:
            feed = feedparser.parse(res.text)
            for entry in feed.entries:
                title_text = entry.get("title", "")
                url = entry.get("link", "")
                desc = clean_html(entry.get("summary", ""))
                posted = entry.get("published", datetime.now().isoformat())
                
                company = "Hacker News Startup"
                title = title_text
                
                # Heuristic to split company name and job title
                for split_phrase in [" is hiring an ", " is hiring a ", " is hiring ", " is looking for ", " wants a ", " wants "]:
                    if split_phrase in title_text:
                        parts = title_text.split(split_phrase, 1)
                        company = parts[0].strip()
                        title = parts[1].strip()
                        break
                        
                jobs.append({
                    "title": title,
                    "company": company,
                    "location": "Remote / USA",
                    "is_remote": True,
                    "source": "hackernews",
                    "url": url,
                    "description": desc,
                    "posted_date": posted
                })
    except Exception as e:
        print(f"[Scraper Error] Hacker News Jobs RSS failed: {e}")
    return jobs

def scrape_reddit_jobs():
    jobs = []
    try:
        print("[Scraper] Fetching Reddit Jobs RSS feeds...")
        import feedparser
        subreddits = ["devops", "sre"]
        for sub in subreddits:
            url = f"https://www.reddit.com/r/{sub}/search.rss?q=flair%3Ahiring+OR+hiring&restrict_sr=1&sort=new"
            res = make_request(url)
            if res:
                feed = feedparser.parse(res.text)
                for entry in feed.entries:
                    title_text = entry.get("title", "")
                    url = entry.get("link", "")
                    desc = clean_html(entry.get("summary", ""))
                    posted = entry.get("published", datetime.now().isoformat())
                    
                    if not any(k in title_text.lower() for k in ["[hiring]", "hiring", "job", "career"]):
                        continue
                        
                    clean_title = title_text
                    for prefix in ["[hiring]", "hiring:", "hiring -", "[hiring/sre]", "[hiring/devops]"]:
                        if clean_title.lower().startswith(prefix):
                            clean_title = clean_title[len(prefix):].strip()
                    
                    company = f"Reddit r/{sub}"
                    
                    jobs.append({
                        "title": clean_title,
                        "company": company,
                        "location": "Remote / Global",
                        "is_remote": True,
                        "source": f"reddit-{sub}",
                        "url": url,
                        "description": desc,
                        "posted_date": posted
                    })
    except Exception as e:
        print(f"[Scraper Error] Reddit RSS failed: {e}")
    return jobs


# ================= SCORING ENGINES =================

def calculate_local_matching(job, profile):
    title = job["title"].lower()
    desc = job["description"].lower()
    location = job["location"].lower()

    # 1. Downrank noise filter
    downrank_kws = profile.get("downrank_keywords", [])
    if any(kw.lower() in title or f" {kw.lower()} " in desc for kw in downrank_kws):
        return {
            "score": 10,
            "matching_skills": [],
            "missing_skills": profile.get("skills", [])[:5],
            "reason": "Procedural Scorer: Downranked due to matching noise or experience requirements."
        }

    # 2. Skills matching
    matching_skills = []
    missing_skills = []
    for skill in profile.get("skills", []):
        pattern = rf"\b{re.escape(skill.lower())}\b"
        if re.search(pattern, desc) or re.search(pattern, title):
            matching_skills.append(skill)
        else:
            missing_skills.append(skill)

    # 3. Location preference bonus
    location_bonus = 0
    if job["is_remote"]:
        location_bonus = 15
    else:
        prefers_loc = any(loc.lower() in location for loc in profile.get("preferred_locations", []))
        if prefers_loc:
            location_bonus = 10

    # 4. Preferred role match
    role_bonus = 0
    prefers_role = any(role.lower() in title for role in profile.get("preferred_roles", []))
    if prefers_role:
        role_bonus = 20

    # 5. Target company bonus
    company_bonus = 0
    target_cos = profile.get("target_companies", [])
    if any(co.lower() in job["company"].lower() for co in target_cos):
        company_bonus = 15

    skill_ratio = len(matching_skills) / len(profile["skills"]) if profile.get("skills") else 0
    base_score = int(skill_ratio * 65)
    final_score = min(100, base_score + role_bonus + location_bonus + company_bonus)
    if final_score < 20:
        final_score = 20

    if final_score >= 80:
        reason = f"Procedural Scorer: Excellent match on {', '.join(matching_skills[:3])}."
    elif final_score >= 60:
        reason = f"Procedural Scorer: Good potential, but missing {', '.join(missing_skills[:2])}."
    else:
        reason = f"Procedural Scorer: Low skill alignment. Missing {', '.join(missing_skills[:3])}."

    return {
        "score": final_score,
        "matching_skills": matching_skills,
        "missing_skills": missing_skills,
        "reason": reason
    }

def score_job_with_gemini(job, profile, api_key):
    if not api_key or api_key == "YOUR_GEMINI_API_KEY" or api_key.strip() == "":
        return calculate_local_matching(job, profile)
        
    try:
        genai.configure(api_key=api_key)
        model = genai.GenerativeModel("gemini-2.5-flash")
        
        prompt = f"""You are an expert job alignment and scoring system. Your task is to evaluate the match between a candidate's profile and a job listing.

Candidate Profile:
- Target Role: {profile.get('title')}
- Candidate Experience: {profile.get('experience_years')} years
- Core Skills: {', '.join(profile.get('skills', []))}
- Preferred Locations: {', '.join(profile.get('preferred_locations', []))}
- Preferred Companies (Boost +15 points): {', '.join(profile.get('target_companies', []))}
- Strong Match Keywords: {', '.join(profile.get('strong_match_keywords', []))}
- Excluded/Noise Keywords (Downrank below 20): {', '.join(profile.get('downrank_keywords', []))}

Job Details:
- Title: {job['title']}
- Company: {job['company']}
- Location: {job['location']}
- Description snippet: {job['description']}

Relevance Scoring Rules:
1. **Experience Check**: The candidate has {profile.get('experience_years')} years of experience. Downrank to a score below 30 only if the job explicitly requires senior, staff, lead, principal titles, or demands 5+ years of experience. Mid-level or 1-4 years experience requirements should NOT be downranked.
2. **Location Check**: If the job is remote, or located in {', '.join(profile.get('preferred_locations', []))}, it is a high-preference fit.
3. **Skills Alignment**: Identify matches between the Core Skills and the Job Description. Be flexible with synonyms (e.g., "AWS" matches "Amazon Web Services", "K8s" matches "Kubernetes", "CI/CD" matches "GitHub Actions/GitOps").
4. **Company Boost**: If the company name matches or contains any of the Preferred Companies, boost the final score by 15 points (capping at 100).
5. **Noise Check**: If the title contains any of the Excluded/Noise Keywords, set the score below 20.

Provide the response in the following JSON format. Return ONLY the JSON object, with no markdown tags (do not wrap in ```json) or additional text.

{{
  "score": <integer from 0 to 100>,
  "matching_skills": [<list of matched skills from Core Skills>],
  "missing_skills": [<list of missing skills from Core Skills that were requested in the description>],
  "reason": "<one sentence summarizing the candidate's alignment and why this score was given>"
}}"""

        response = model.generate_content(prompt)
        text = response.text.strip()
        if text.startswith("```"):
            text = text.replace("```json", "", 1).replace("```", "", 1).strip()
            
        data = json.loads(text)
        return {
            "score": int(data.get("score", 50)),
            "matching_skills": data.get("matching_skills", []),
            "missing_skills": data.get("missing_skills", []),
            "reason": data.get("reason", "AI evaluation completed successfully.")
        }
    except Exception as e:
        print(f"[Gemini Error] AI scoring failed, falling back to local engine: {e}")
        return calculate_local_matching(job, profile)

# ================= GOOGLE SHEETS =================

def write_to_google_sheet(jobs_list, sheets_config):
    json_key_path = os.path.join(BASE_DIR, sheets_config.get("service_account_json", "service_account.json"))
    sheet_id = sheets_config.get("sheet_id")
    
    if not sheet_id or sheet_id == "YOUR_GOOGLE_SHEET_ID" or not os.path.exists(json_key_path):
        print("[Sheets] Google Sheets config or service account file missing. Skipping Google Sheets upload.")
        return

    try:
        print(f"[Sheets] Authenticating with Google Sheets via {json_key_path}...")
        scopes = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive"
        ]
        creds = Credentials.from_service_account_file(json_key_path, scopes=scopes)
        client = gspread.authorize(creds)
        
        print(f"[Sheets] Opening spreadsheet by ID: {sheet_id}...")
        spreadsheet = client.open_by_key(sheet_id)
        sheet = spreadsheet.get_worksheet(0)
        
        existing_values = sheet.get_all_values()
        headers = ["Score", "Title", "Company", "Location", "Source", "URL", "Reason", "Matching Skills", "Missing Skills", "Posted Date", "Date Aggregated"]
        
        if not existing_values:
            sheet.append_row(headers)
            
        rows_to_append = []
        for j in jobs_list:
            rows_to_append.append([
                j["score"],
                j["title"],
                j["company"],
                j["location"],
                j["source"],
                j["url"],
                j["reason"],
                j["matching_skills"],
                j["missing_skills"],
                j["posted_date"],
                datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ])
            
        if rows_to_append:
            sheet.append_rows(rows_to_append)
            print(f"[Sheets] Successfully appended {len(rows_to_append)} rows to Google Sheet!")
    except Exception as e:
        print(f"[Sheets Error] Failed to write to Google Sheets: {e}")

# ================= EMAIL INGEST =================

def send_digest_email(jobs_list, excel_path, config, scraper_config, total_scraped):
    email_config = config["email"]
    sender = email_config["sender_email"]
    password = email_config["sender_password"]
    receiver = email_config["receiver_email"]
    smtp_user = email_config.get("smtp_username", sender)
    
    if sender == "sender@gmail.com" or password == "YOUR_GMAIL_APP_PASSWORD" or password == "YOUR_SMTP2GO_PASSWORD":
        print("[Mailer] Email credentials not configured. Skipping email delivery.")
        return

    # Extract dynamic stats
    date_pretty = datetime.now().strftime("%b %d, %Y")
    total_new = len(jobs_list)
    
    # Calculate matches categories
    strong_matches = [j for j in jobs_list if j["score"] >= 80]
    strong_matches = sorted(strong_matches, key=lambda x: x["score"], reverse=True)[:5]
    
    summary_matches = [j for j in jobs_list if 70 <= j["score"] <= 79]
    summary_matches = sorted(summary_matches, key=lambda x: x["score"], reverse=True)
    
    strong_matches_count = len([j for j in jobs_list if j["score"] >= 80])

    # Google sheet link
    sheets_config = config.get("google_sheets", {})
    sheet_id = sheets_config.get("sheet_id", "")
    sheet_link = f"https://docs.google.com/spreadsheets/d/{sheet_id}/edit" if sheet_id and sheet_id != "YOUR_GOOGLE_SHEET_ID" else "#"

    msg = MIMEMultipart()
    msg["From"] = sender
    msg["To"] = receiver
    msg["Subject"] = f"Daily DevOps Job Digest — {date_pretty} — {total_new} new matches"

    # Build Top 5 HTML Cards
    top_matches_html = ""
    if strong_matches:
        for j in strong_matches:
            # Score badge color: >=90 gold (#C9A84C), 80-89 deep red (#8B1A1A)
            badge_color = "#C9A84C" if j["score"] >= 90 else "#8B1A1A"
            top_matches_html += f"""
            <div style="padding: 20px; border: 1px solid rgba(201, 168, 76, 0.4); border-radius: 6px; margin-bottom: 20px; background-color: #ffffff; text-align: left;">
              <table width="100%" cellpadding="0" cellspacing="0" border="0" style="text-align: left;">
                <tr>
                  <td style="text-align: left; vertical-align: top; padding-right: 10px;">
                    <span style="font-weight: bold; font-size: 16px; color: #1A1A2E;">{j['company']}</span>
                    <span style="font-size: 16px; color: #1A1A2E;"> - {j['title']}</span>
                  </td>
                  <td align="right" valign="top" style="width: 60px; text-align: right; vertical-align: top;">
                    <span style="background-color: {badge_color}; color: #ffffff; font-weight: bold; font-size: 12px; padding: 5px 10px; border-radius: 4px; font-family: monospace; display: inline-block; text-align: center; min-width: 40px;">
                      {j['score']}%
                    </span>
                  </td>
                </tr>
              </table>
              <p style="margin: 10px 0; font-size: 12px; color: #1A1A2E; opacity: 0.8; text-align: left;">
                📍 {j['location']} &middot; 📂 {j['source'].upper()} &middot; 📅 {j['posted_date']}
              </p>
              <p style="margin: 8px 0 16px 0; font-size: 13px; font-style: italic; color: #1A1A2E; text-align: left; line-height: 1.4;">
                &ldquo;{j['reason']}&rdquo;
              </p>
              <a href="{j['url']}" target="_blank" style="background-color: #C9A84C; color: #1A1A2E; padding: 10px 20px; border-radius: 4px; text-decoration: none; font-size: 12px; font-weight: bold; display: inline-block;">
                View Job
              </a>
            </div>
            """
    else:
        top_matches_html = """
        <p style="font-size: 13px; color: #1A1A2E; opacity: 0.7; font-style: italic; text-align: center; padding: 24px; border: 1px dashed rgba(201, 168, 76, 0.4); border-radius: 6px; margin-bottom: 20px;">
          No strong matches (&ge;80%) today.
        </p>
        """

    # Build Summary Table HTML
    summary_table_html = ""
    if summary_matches:
        summary_rows = ""
        for j in summary_matches:
            # Score badge 70-79: #1B4332 (dark green)
            badge_color = "#1B4332"
            summary_rows += f"""
            <tr style="border-bottom: 1px solid rgba(201, 168, 76, 0.4); text-align: left;">
              <td style="border: 1px solid rgba(201, 168, 76, 0.4); padding: 10px; font-weight: bold; color: #1A1A2E; text-align: left;">{j['company']}</td>
              <td style="border: 1px solid rgba(201, 168, 76, 0.4); padding: 10px; color: #1A1A2E; text-align: left;">{j['title']}</td>
              <td style="border: 1px solid rgba(201, 168, 76, 0.4); padding: 10px; text-align: center;">
                <span style="background-color: {badge_color}; color: #ffffff; font-weight: bold; font-size: 11px; padding: 3px 6px; border-radius: 3px; font-family: monospace; display: inline-block;">
                  {j['score']}%
                </span>
              </td>
              <td style="border: 1px solid rgba(201, 168, 76, 0.4); padding: 10px; color: #1A1A2E; text-align: left;">{j['location']}</td>
              <td style="border: 1px solid rgba(201, 168, 76, 0.4); padding: 10px; text-align: center;">
                <a href="{j['url']}" target="_blank" style="color: #C9A84C; font-weight: bold; text-decoration: underline;">View</a>
              </td>
            </tr>
            """
        summary_table_html = f"""
        <table width="100%" cellpadding="0" cellspacing="0" style="font-size: 13px; color: #1A1A2E; border-collapse: collapse; border: 1px solid rgba(201, 168, 76, 0.4); text-align: left;">
          <thead>
            <tr style="background-color: #1A1A2E; text-align: left; font-weight: bold; color: #ffffff;">
              <th style="border: 1px solid rgba(201, 168, 76, 0.4); padding: 12px; color: #ffffff;">Company</th>
              <th style="border: 1px solid rgba(201, 168, 76, 0.4); padding: 12px; color: #ffffff;">Role</th>
              <th style="border: 1px solid rgba(201, 168, 76, 0.4); padding: 12px; text-align: center; width: 60px; color: #ffffff;">Score</th>
              <th style="border: 1px solid rgba(201, 168, 76, 0.4); padding: 12px; color: #ffffff;">Location</th>
              <th style="border: 1px solid rgba(201, 168, 76, 0.4); padding: 12px; text-align: center; width: 50px; color: #ffffff;">Link</th>
            </tr>
          </thead>
          <tbody>
            {summary_rows}
          </tbody>
        </table>
        """
    else:
        summary_table_html = """
        <p style="font-size: 13px; color: #1A1A2E; opacity: 0.7; font-style: italic; text-align: center; padding: 24px; border: 1px dashed rgba(201, 168, 76, 0.4); border-radius: 6px;">
          No medium recommendation fits (70-79) today.
        </p>
        """

    html_body = f"""
    <html>
      <body style="margin: 0; padding: 24px; font-family: system-ui, Arial, sans-serif; background-color: #ffffff; color: #1A1A2E; text-align: left; -webkit-font-smoothing: antialiased;">
        <div style="max-width: 600px; margin: 0 auto; background-color: #ffffff; border: 1px solid rgba(201, 168, 76, 0.4); border-radius: 8px; overflow: hidden; text-align: left;">
          
          <!-- 1. Header section with Gold Stripe -->
          <div style="background-color: #1A1A2E; padding: 28px 24px; text-align: center; color: #ffffff;">
            <h1 style="margin: 0; font-size: 26px; font-weight: bold; color: #ffffff; letter-spacing: -0.5px;">Your Daily Job Digest</h1>
            <p style="margin: 10px 0 0 0; font-size: 14px; color: #C9A84C; font-weight: 500;">
              {date_pretty} &middot; {total_new} new jobs &middot; {strong_matches_count} strong matches (&ge;80%)
            </p>
          </div>
          <div style="height: 4px; background-color: #C9A84C;"></div>
          
          <div style="padding: 28px 24px; text-align: left;">
            <!-- 2. Top 5 Matches section -->
            <h2 style="font-size: 13px; font-weight: bold; margin-top: 0; margin-bottom: 20px; color: #1A1A2E; border-bottom: 2px solid rgba(201, 168, 76, 0.2); padding-bottom: 8px; font-family: monospace; text-transform: uppercase; letter-spacing: 1px; text-align: left;">🎯 Top Recommendation Fits (Score &ge; 80%)</h2>
            {top_matches_html}
            
            <!-- 3. Summary table below top 5 -->
            <h2 style="font-size: 13px; font-weight: bold; margin-top: 40px; margin-bottom: 20px; color: #1A1A2E; border-bottom: 2px solid rgba(201, 168, 76, 0.2); padding-bottom: 8px; font-family: monospace; text-transform: uppercase; letter-spacing: 1px; text-align: left;">📋 Medium Recommendation Fits (Score 70-79)</h2>
            {summary_table_html}
          </div>
          
          <!-- 4. Stats footer & 5. Footer -->
          <div style="background-color: #F8F5EE; border-top: 1px solid rgba(201, 168, 76, 0.4); padding: 28px 24px; color: #1A1A2E; font-size: 12px; line-height: 1.6; text-align: left;">
            <h3 style="font-size: 11px; font-weight: bold; color: #1A1A2E; margin: 0 0 10px 0; text-transform: uppercase; letter-spacing: 1px; font-family: monospace; text-align: left;">📊 DIGEST METRICS</h3>
            <p style="margin: 0; text-align: left;">Sources checked: 7</p>
            <p style="margin: 0; text-align: left;">Total scraped: {total_scraped}</p>
            <p style="margin: 0; text-align: left;">After filtering: {total_new}</p>
            <p style="margin: 0; text-align: left;">Strong matches (&ge;80%): {strong_matches_count}</p>
            <p style="margin: 14px 0 0 0; font-weight: bold; text-align: left;">
              🔗 <a href="{sheet_link}" target="_blank" style="color: #1A1A2E; text-decoration: underline;">View full sheet (Google Sheets)</a>
            </p>
            
            <hr style="border: 0; border-top: 1px solid rgba(201, 168, 76, 0.2); margin: 24px 0;">
            
            <p style="margin: 0; text-align: center; font-size: 11px; color: #1A1A2E; opacity: 0.8;">
              This digest was generated automatically. &middot; <a href="#" style="color: #1A1A2E; text-decoration: underline;">Unsubscribe</a>
            </p>
          </div>
          
        </div>
      </body>
    </html>
    """

    msg.attach(MIMEText(html_body, "html"))

    # Attach Excel sheet
    try:
        with open(excel_path, "rb") as attachment:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header(
                "Content-Disposition",
                f"attachment; filename= {os.path.basename(excel_path)}",
            )
            msg.attach(part)
    except Exception as e:
        print(f"[Mailer Error] Failed to attach Excel sheet: {e}")
        return

    # SMTP Send
    try:
        print(f"[Mailer] Connecting to SMTP server {email_config['smtp_server']}:{email_config['smtp_port']}...")
        server = smtplib.SMTP(email_config["smtp_server"], email_config["smtp_port"])
        server.starttls()
        server.login(smtp_user, password)
        server.sendmail(sender, receiver, msg.as_string())
        server.quit()
        print("[Mailer] Digest email sent successfully!")
    except Exception as e:
        print(f"[Mailer Error] Failed to send email via SMTP: {e}")

def style_excel_sheet(excel_path):
    try:
        print(f"[Compiler] Styling Excel sheet: {excel_path}...")
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        
        # Set sheet tab color
        ws.sheet_properties.tabColor = "C9A84C"
        
        # Fills & Fonts definitions
        header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        
        gold_fill = PatternFill(start_color="FFF8E7", end_color="FFF8E7", fill_type="solid")
        red_fill = PatternFill(start_color="FFF0F0", end_color="FFF0F0", fill_type="solid")
        green_fill = PatternFill(start_color="F0FFF4", end_color="F0FFF4", fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        alt_fill = PatternFill(start_color="FAFAF7", end_color="FAFAF7", fill_type="solid")
        
        # Border definition: #C9A84C
        thin_border_side = Side(style='thin', color='C9A84C')
        row_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
        
        # 1. Style Header Row
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = row_border
            
        # 2. Style Data Rows based on Score
        for row_idx in range(2, ws.max_row + 1):
            score_cell = ws.cell(row=row_idx, column=1)
            score = 0
            try:
                score = int(score_cell.value)
            except Exception:
                pass
                
            # Determine fill
            if score >= 90:
                row_fill = gold_fill
            elif score >= 80:
                row_fill = red_fill
            elif score >= 70:
                row_fill = green_fill
            else:
                # Alternating row subtle tint
                row_fill = alt_fill if row_idx % 2 == 1 else white_fill
                
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = row_fill
                cell.font = Font(name="Calibri", size=10, color="1A1A2E")
                cell.border = row_border
                
                # Align score column to center, link to left, etc.
                if col_idx in [1, 5, 10]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    
        # Auto-adjust column widths
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            # Limit width to maximum 45
            ws.column_dimensions[col_letter].width = min(45, max(max_len + 3, 10))
            
        wb.save(excel_path)
        print("[Compiler] Excel styled successfully!")
    except Exception as e:
        print(f"[Compiler Error] Failed to style Excel sheet: {e}")

# ================= MAIN ORCHESTRATOR =================

def main():
    print(f"=== OpsHunt AI CLI Digest Booted at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} ===")
    
    try:
        config = load_config()
    except Exception as e:
        print(f"[Critical Error] Failed to load config: {e}")
        return

    profile = config["candidate_profile"]
    api_key = config["gemini_api_key"]
    scraper_config = config.get("scraper", {})
    seen_jobs = load_seen_jobs()

    # 1. Scrape all sources
    raw_jobs = []
    
    remotive_jobs = scrape_remotive()
    jobicy_jobs = scrape_jobicy()
    wwr_jobs = scrape_wwr()
    otta_jobs = scrape_otta()
    arbeitnow_jobs = scrape_arbeitnow()
    naukri_jobs = scrape_naukri()
    instahyre_jobs = scrape_instahyre()
    jobspy_jobs = scrape_jobspy()
    remoteok_jobs = scrape_remoteok()
    hn_jobs = scrape_hn_jobs()
    reddit_jobs = scrape_reddit_jobs()
    
    raw_jobs.extend(remotive_jobs)
    raw_jobs.extend(jobicy_jobs)
    raw_jobs.extend(wwr_jobs)
    raw_jobs.extend(otta_jobs)
    raw_jobs.extend(arbeitnow_jobs)
    raw_jobs.extend(naukri_jobs)
    raw_jobs.extend(instahyre_jobs)
    raw_jobs.extend(jobspy_jobs)
    raw_jobs.extend(remoteok_jobs)
    raw_jobs.extend(hn_jobs)
    raw_jobs.extend(reddit_jobs)
    
    # Store source counts
    source_fetched = {
        "remotive": len(remotive_jobs),
        "jobicy": len(jobicy_jobs),
        "weworkremotely": len(wwr_jobs),
        "otta": len(otta_jobs),
        "arbeitnow": len(arbeitnow_jobs),
        "naukri": len(naukri_jobs),
        "instahyre": len(instahyre_jobs),
        "jobspy": len(jobspy_jobs),
        "remoteok": len(remoteok_jobs),
        "hackernews": len(hn_jobs),
        "reddit": len(reddit_jobs)
    }
    
    source_passed = {k: 0 for k in source_fetched.keys()}
    
    print(f"[Scraper] Finished scraping. Total raw jobs parsed: {len(raw_jobs)}")

    # 2. Filter, Deduplicate and Score
    new_jobs = []
    filtered_out_by_age = 0
    filtered_out_by_role = 0
    filtered_out_by_noise = 0
    filtered_out_by_dead_link = 0
    filtered_out_by_duplicate = 0
    
    max_age_days = scraper_config.get("max_job_age_days", 7)
    unique_job_keys = set()

    for job in raw_jobs:
        source = job["source"]
        
        # Prevent in-memory duplicates
        job_key = f"{job['company'].lower()}::{job['title'].lower()}"
        if job_key in unique_job_keys:
            filtered_out_by_duplicate += 1
            continue
        unique_job_keys.add(job_key)

        # Check job age
        if job.get("posted_date"):
            try:
                posted_dt = pd.to_datetime(job["posted_date"])
                age_days = (datetime.now() - posted_dt.replace(tzinfo=None)).days
                if age_days > max_age_days:
                    filtered_out_by_age += 1
                    continue
            except Exception:
                pass

        # Check against previously emailed seen jobs list
        if job_url := job.get("url"):
            clean_url = job_url.split("?")[0]
            job["url"] = clean_url
            
            if clean_url in seen_jobs:
                filtered_out_by_duplicate += 1
                continue

            # Filter: title must contain at least one of the TITLE_KEYWORDS (case insensitive)
            TITLE_KEYWORDS = [
                "devops", "sre", "reliability", "platform",
                "cloud", "infrastructure", "kubernetes", "k8s",
                "devsecops", "gitops", "mlops", "site reliability"
            ]
            title_lower = job["title"].lower()
            roles_match = any(kw in title_lower for kw in TITLE_KEYWORDS)
            if not roles_match:
                filtered_out_by_role += 1
                continue

            # Filter out: titles containing any downrank_keywords
            if any(kw.lower() in title_lower for kw in profile.get("downrank_keywords", [])):
                filtered_out_by_noise += 1
                continue

            # Check url liveness
            if not is_url_alive(clean_url):
                filtered_out_by_dead_link += 1
                continue

            # Truncate description to first 1500 chars for AI scoring
            job_clean = job.copy()
            job_clean["description"] = job["description"][:1500]
            job_clean["url"] = clean_url
            
            new_jobs.append(job_clean)
            matched_key = None
            if source in source_passed:
                matched_key = source
            else:
                for k in source_passed.keys():
                    if source.startswith(k):
                        matched_key = k
                        break
            if matched_key:
                source_passed[matched_key] += 1

    # Log: how many jobs fetched per source, how many after filtering.
    print("\n=== SCRAPING & FILTERING METRICS ===")
    for src in source_fetched.keys():
        print(f" - {src.upper():<15} | Fetched: {source_fetched[src]:>3} | Passed: {source_passed[src]:>3}")
    print(f"Total raw jobs: {len(raw_jobs)}")
    print(f"Skipped: {filtered_out_by_age} older than {max_age_days}d, {filtered_out_by_role} role mismatch, {filtered_out_by_noise} noise keywords, {filtered_out_by_dead_link} dead links, {filtered_out_by_duplicate} duplicates")
    print(f"Final net jobs to score: {len(new_jobs)}\n")

    if not new_jobs:
        print("[Process] No new jobs found to process today. Exiting.")
        return

    # 3. AI Score jobs
    import time
    scored_jobs = []
    print("[AI Evaluator] Initiating matching evaluations...")
    for idx, job in enumerate(new_jobs):
        # Count strong match keywords in title and description
        strong_kws = profile.get("strong_match_keywords", [])
        desc_lower = job.get("description", "").lower()
        title_lower = job.get("title", "").lower()
        
        match_count = 0
        matched_kws = []
        for kw in strong_kws:
            pattern = rf"\b{re.escape(kw.lower())}\b"
            if re.search(pattern, desc_lower) or re.search(pattern, title_lower):
                match_count += 1
                matched_kws.append(kw)
        
        # Decide if we score with Gemini
        if match_count >= 2:
            # We wait 4s to stay within rate limits if calling Gemini
            # Only sleep if we have already called Gemini for at least one job
            gemini_called_count = len([j for j in scored_jobs if "Skipped AI scoring" not in j.get("reason", "")])
            if gemini_called_count > 0 and api_key and api_key != "YOUR_GEMINI_API_KEY" and api_key.strip() != "":
                time.sleep(4.0)
            
            print(f"[AI Evaluator] ({idx+1}/{len(new_jobs)}) Scoring: '{job['title']}' at '{job['company']}' (matches: {match_count} - {', '.join(matched_kws)})")
            score_res = score_job_with_gemini(job, profile, api_key)
        else:
            print(f"[AI Evaluator] ({idx+1}/{len(new_jobs)}) Skipping Gemini for: '{job['title']}' at '{job['company']}' (only {match_count} strong match keywords: {', '.join(matched_kws) if matched_kws else 'none'}). Assigned score 50.")
            # Run local matching to populate matched/missing skills
            local_res = calculate_local_matching(job, profile)
            score_res = {
                "score": 50,
                "matching_skills": local_res["matching_skills"],
                "missing_skills": local_res["missing_skills"],
                "reason": f"Skipped AI scoring: Only {match_count} strong match keywords found ({', '.join(matched_kws) if matched_kws else 'none'})."
            }
        
        job_scored = job.copy()
        job_scored["score"] = score_res["score"]
        job_scored["matching_skills"] = ", ".join(score_res["matching_skills"])
        job_scored["missing_skills"] = ", ".join(score_res["missing_skills"])
        job_scored["reason"] = score_res["reason"]
        scored_jobs.append(job_scored)

    # 4. Save to Excel
    df = pd.DataFrame(scored_jobs)
    columns_order = ["score", "title", "company", "location", "source", "url", "reason", "matching_skills", "missing_skills", "posted_date"]
    df = df[columns_order]
    df = df.sort_values(by="score", ascending=False)

    date_str = datetime.now().strftime("%Y-%m-%d")
    excel_filename = f"DevOps_Jobs_Digest_{date_str}.xlsx"
    excel_path = os.path.join(BASE_DIR, excel_filename)

    print(f"[Compiler] Writing results to Excel sheet: {excel_filename}...")
    df.to_excel(excel_path, index=False, sheet_name="Jobs Digest")
    style_excel_sheet(excel_path)

    # 5. Write to Google Sheets
    if "google_sheets" in config:
        write_to_google_sheet(scored_jobs, config["google_sheets"])

    # 6. Email Digest
    send_digest_email(scored_jobs, excel_path, config, scraper_config, len(raw_jobs))

    # 7. Save tracked URLs to avoid duplicate emails tomorrow
    for job in new_jobs:
        seen_jobs.add(job["url"])
    save_seen_jobs(seen_jobs)
    
    print("=== OpsHunt AI CLI Digest Execution Completed Successfully ===")

if __name__ == "__main__":
    main()
